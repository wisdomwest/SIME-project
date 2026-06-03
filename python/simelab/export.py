"""
export.py — CSV / Excel export for all analysis results.

Usage:
    from export import ExportManager
    em = ExportManager(output_dir="./exports")
    em.export_all(G, fe, sa, da, ha, ca)
    # Writes: features.csv, sentiment.csv, disinfo.csv, hashtag.csv, censorship.csv, full_report.xlsx
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Optional

import networkx as nx


class ExportManager:
    """
    Export analysis results to CSV and Excel files.
    """

    def __init__(self, output_dir: str = "./exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def _path(self, name: str, ext: str = "csv") -> str:
        return str(self.output_dir / f"{name}_{self.timestamp}.{ext}")

    def export_features_csv(self, fe) -> str:
        """Export 9-D feature matrix as CSV."""
        path = self._path("features")
        df = fe.to_dataframe()
        df.to_csv(path)
        return path

    def export_sentiment_csv(self, sa) -> str:
        """Export sentiment labels as CSV."""
        path = self._path("sentiment")
        df = sa.to_dataframe()
        df.to_csv(path)
        return path

    def export_disinformation_csv(self, da) -> str:
        """Export disinformation scores as CSV."""
        path = self._path("disinformation")
        df = da.to_dataframe()
        df.to_csv(path)
        return path

    def export_hashtag_csv(self, ha) -> str:
        """Export hashtag analysis as CSV."""
        path = self._path("hashtag")
        df = ha.to_dataframe()
        df.to_csv(path)
        return path

    def export_censorship_csv(self, ca) -> str:
        """Export structural hole analysis as CSV."""
        path = self._path("censorship")
        df = ca.to_dataframe()
        df.to_csv(path)
        return path

    def export_full_excel(self, G, fe, sa=None, da=None, ha=None, ca=None) -> str:
        """
        Export all analyses to a single multi-sheet Excel workbook.

        Sheets:
          - Graph Summary (basic stats)
          - Features (9-D matrix)
          - Sentiment (if sa provided)
          - Disinformation (if da provided)
          - Hashtags (if ha provided)
          - Censorship (if ca provided)
          - Top Influencers
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        path = self._path("full_report", "xlsx")
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Header style
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        def write_df(ws, df, title=None):
            """Write a DataFrame to a worksheet with styled headers."""
            if title:
                ws.append([title])
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) + 1)
                ws.cell(1, 1).font = Font(bold=True, size=14, color="1B5E20")

            # Headers
            headers = [df.index.name or "index"] + list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=(3 if title else 1), column=col_idx, value=header)
                cell.font = header_font_white
                cell.fill = header_fill

            # Data
            start_row = 4 if title else 2
            for row_idx, (idx, row) in enumerate(df.iterrows()):
                ws.cell(row=start_row + row_idx, column=1, value=str(idx))
                for col_idx, val in enumerate(row, 2):
                    ws.cell(row=start_row + row_idx, column=col_idx, value=val)

        # Sheet 1: Graph Summary
        ws_summary = wb.create_sheet("Graph Summary")
        ws_summary.append(["SIMElab Data Explorer — Analysis Report"])
        ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_summary.append([])
        ws_summary.append(["Nodes", G.number_of_nodes()])
        ws_summary.append(["Edges", G.number_of_edges()])

        if G.is_directed():
            ws_summary.append(["Directed", "Yes"])
            ws_summary.append(["Reciprocity", f"{nx.overall_reciprocity(G):.4f}"])
        else:
            ws_summary.append(["Directed", "No"])

        import networkx as nx
        ws_summary.append(["Density", f"{nx.density(G):.6f}"])
        ws_summary.append(["Connected Components",
                           nx.number_weakly_connected_components(G) if G.is_directed()
                           else nx.number_connected_components(G)])

        # Sheet 2: Features
        ws_feat = wb.create_sheet("Features")
        write_df(ws_feat, fe.to_dataframe(), "9-Dimensional Feature Vectors")

        # Sheet 3: Top Influencers
        ws_inf = wb.create_sheet("Top Influencers")
        from features import INFLUENCE
        top_nodes = fe.get_top(INFLUENCE, k=20)
        ws_inf.append(["Top 20 Influencers"])
        ws_inf.append([])
        ws_inf.append(["Rank", "Username", "Display Name", "Influence Score", "Followers"])
        for rank, (node, score) in enumerate(top_nodes, 1):
            name = G.nodes[node].get("display_name", "")
            followers = G.nodes[node].get("followers", "")
            ws_inf.append([rank, node, name, round(score, 6), followers])

        # Sheet 4-7: Optional analyses
        if sa is not None:
            ws_sent = wb.create_sheet("Sentiment")
            write_df(ws_sent, sa.to_dataframe(), "K-Means Sentiment Clustering")

        if da is not None:
            ws_dis = wb.create_sheet("Disinformation")
            write_df(ws_dis, da.to_dataframe(), "5-Signal Disinformation Scores")

        if ha is not None:
            ws_hash = wb.create_sheet("Hashtags")
            write_df(ws_hash, ha.to_dataframe(), "Hashtag Lifecycle & Authenticity")

        if ca is not None:
            ws_cen = wb.create_sheet("Censorship")
            write_df(ws_cen, ca.to_dataframe(), "Structural Hole Analysis")

        wb.save(path)
        return path

    def export_all(self, G, fe, sa=None, da=None, ha=None, ca=None) -> dict:
        """
        Export all analyses. Returns dict of {name: filepath}.
        """
        results = {}

        # Individual CSVs
        results["features"] = self.export_features_csv(fe)
        if sa:
            results["sentiment"] = self.export_sentiment_csv(sa)
        if da:
            results["disinformation"] = self.export_disinformation_csv(da)
        if ha:
            results["hashtag"] = self.export_hashtag_csv(ha)
        if ca:
            results["censorship"] = self.export_censorship_csv(ca)

        # Combined Excel
        results["full_report"] = self.export_full_excel(G, fe, sa, da, ha, ca)

        return results


# ─── CLI ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    from loader import load_nodexl
    from features import FeatureEngineer

    if len(sys.argv) < 2:
        print("Usage: python export.py <file.xlsx|file.csv>")
        sys.exit(1)

    filepath = sys.argv[1]
    G, meta = load_nodexl(filepath)
    fe = FeatureEngineer(G)
    fe.build_matrix()

    em = ExportManager(output_dir="./exports")
    results = em.export_all(G, fe)
    for name, path in results.items():
        print(f"Exported {name}: {path}")
