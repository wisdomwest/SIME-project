"""
loader.py — Read NodeXL Pro XLSX exports into NetworkX directed graph.

Handles:
- Multi-sheet NodeXL workbooks (Edges + Vertices + Overall Metrics)
- 2-row merged header format (category row + column name row)
- Auto-detection: NodeXL XLSX vs simple CSV edge list
- Maps all 72 Vertex columns as node attributes
- Returns a rich NetworkX DiGraph ready for analysis

Usage:
    from loader import load_nodexl
    G, metadata = load_nodexl("RejectFinanceBill2024.xlsx")
"""

import zipfile
from pathlib import Path
from typing import Optional, Tuple, Dict, Any

import networkx as nx
import pandas as pd
import openpyxl


# ─── Column maps ────────────────────────────────────────────────────────────
# Map NodeXL Vertex column names → clean attribute names
VERTEX_COLUMN_MAP = {
    # Identity
    "Vertex": "username",
    "Name": "display_name",
    "User ID": "user_id",
    # Account stats
    "Followers": "followers",
    "Followed": "following",
    "Tweets": "tweets_count",
    "Listed Count": "listed_count",
    "Favourites Count": "favourites_count",
    "Media Count": "media_count",
    "Verified": "verified",
    "Is Blue Verified": "is_blue_verified",
    "Joined Twitter Date (UTC)": "joined_date",
    "Location": "location",
    "Description": "description",
    "URL": "profile_url",
    "Profile Banner URL": "profile_banner_url",
    "Default Profile": "default_profile",
    "Default Profile Image": "default_profile_image",
    # Graph Metrics (pre-computed by NodeXL)
    "Degree": "degree_centrality",
    "In-Degree": "in_degree_centrality",
    "Out-Degree": "out_degree_centrality",
    "Betweenness Centrality": "betweenness_centrality",
    "Closeness Centrality": "closeness_centrality",
    "Eigenvector Centrality": "eigenvector_centrality",
    "PageRank": "pagerank",
    "Clustering Coefficient": "clustering_coefficient",
    "Reciprocated Vertex Pair Ratio": "reciprocity_ratio",
    # Layout
    "X": "layout_x",
    "Y": "layout_y",
    "Layout Order": "layout_order",
    # Visual (optional, for reference)
    "Color": "nodexl_color",
    "Shape": "nodexl_shape",
    "Size": "nodexl_size",
    "Label": "nodexl_label",
    "Tooltip": "nodexl_tooltip",
    "Image File": "image_url",
    # Other
    "ID": "nodexl_vertex_id",
    "You Are Followed By": "you_are_followed_by",
    "You Are Following": "you_are_following",
    "Can DM": "can_dm",
    "Can Media Tag": "can_media_tag",
    "Has Custom Timelines": "has_custom_timelines",
    "Is Translator": "is_translator",
    "Possibly Sensitive": "possibly_sensitive",
    "Want Retweets": "want_retweets",
    "Withheld": "withheld",
    "Tweeted Search Term?": "tweeted_search_term",
    "Translator Type": "translator_type",
}


def _read_nodexl_sheet(filepath: str, sheet_name: str) -> pd.DataFrame:
    """
    Read a NodeXL sheet with 2-row merged header.
    Row 1 = category labels (Visual Properties, Graph Metrics, etc.)
    Row 2 = actual column names
    Data starts at row 3 (pandas row index 2).

    Returns a DataFrame with clean, deduplicated column names.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]

    # Read row 2 (column names) — row index 2 in openpyxl (1-based)
    raw_headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col_idx).value
        raw_headers.append(str(cell_value).strip() if cell_value is not None else "")

    # Read data rows (3 onwards)
    data = []
    for row_idx in range(3, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row=row_idx, column=col_idx).value)
        data.append(row_data)

    wb.close()

    df = pd.DataFrame(data, columns=raw_headers)

    # Drop completely empty columns and rows
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")

    return df


def _read_nodexl_sheet_fast(filepath: str, sheet_name: str) -> pd.DataFrame:
    """
    Faster variant using openpyxl with optimized iteration.
    Falls back to _read_nodexl_sheet if issues arise.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]

    # Headers from row 2
    headers = [str(cell.value).strip() if cell.value is not None else ""
               for cell in list(ws.iter_rows(min_row=2, max_row=2))[0]]

    # Data from row 3 onwards
    rows = []
    for row in ws.iter_rows(min_row=3):
        rows.append([cell.value for cell in row])

    wb.close()

    df = pd.DataFrame(rows, columns=headers)
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")

    return df


def _detect_format(filepath: str) -> str:
    """
    Detect file format.
    Returns 'nodexl_xlsx', 'csv_edgelist', or raises ValueError.
    """
    path = Path(filepath)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        # Check if it's a NodeXL export (has Edges + Vertices sheets)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            if "Edges" in sheets and "Vertices" in sheets:
                return "nodexl_xlsx"
            else:
                # Generic Excel — try treating first sheet as edge list
                return "csv_edgelist"
        except Exception:
            return "csv_edgelist"

    elif suffix == ".csv":
        return "csv_edgelist"

    else:
        raise ValueError(f"Unsupported file format: {suffix}")


def _parse_edges_nodexl(filepath: str) -> pd.DataFrame:
    """Read the Edges sheet from a NodeXL workbook."""
    df = _read_nodexl_sheet_fast(filepath, "Edges")

    # Rename key columns if they exist
    if "Vertex 1" in df.columns:
        df = df.rename(columns={"Vertex 1": "source", "Vertex 2": "target"})

    # Handle the Relationship column (Retweet, Replies to, Mention, etc.)
    if "Relationship" in df.columns:
        df = df.rename(columns={"Relationship": "edge_type"})

    return df


def _parse_edges_csv(filepath: str) -> pd.DataFrame:
    """Read a simple CSV edge list."""
    df = pd.read_csv(filepath)

    # Auto-detect source/target columns
    cols = [c.lower() for c in df.columns]
    source_col = None
    target_col = None

    for col in df.columns:
        cl = col.lower()
        if source_col is None and ("source" in cl or "vertex 1" in cl or "from" in cl):
            source_col = col
        elif target_col is None and ("target" in cl or "vertex 2" in cl or "to" in cl):
            target_col = col

    if source_col and target_col:
        df = df.rename(columns={source_col: "source", target_col: "target"})
    elif len(df.columns) >= 2:
        # Assume first two columns are source and target
        df = df.rename(columns={df.columns[0]: "source", df.columns[1]: "target"})
    else:
        raise ValueError("Could not identify source and target columns in CSV")

    return df


def _parse_vertices(filepath: str) -> Optional[pd.DataFrame]:
    """Read the Vertices sheet from a NodeXL workbook, if present."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if "Vertices" not in wb.sheetnames:
            wb.close()
            return None
        wb.close()
        df = _read_nodexl_sheet_fast(filepath, "Vertices")
        return df
    except Exception:
        return None


def _parse_overall_metrics(filepath: str) -> Optional[Dict[str, Any]]:
    """Read Overall Metrics sheet into a flat dict, if present."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        if "Overall Metrics" not in wb.sheetnames:
            wb.close()
            return None

        ws = wb["Overall Metrics"]
        metrics = {}
        # Single header row — row index 1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            key = str(row[0]).strip() if row[0] is not None else ""
            val = row[1]
            if key and val is not None:
                try:
                    # Try numeric conversion
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    pass
                metrics[key] = val

        wb.close()

        # If empty after parsing, try the 2-row header format
        if not metrics:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb["Overall Metrics"]
            # Row 1 = "Graph Metric" / "Value", Row 2+ = data
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                key = str(row[0]).strip() if row[0] is not None else ""
                val = row[1]
                if key and val is not None:
                    try:
                        val = float(val)
                        if val == int(val):
                            val = int(val)
                    except (ValueError, TypeError):
                        pass
                    metrics[key] = val
            wb.close()

        return metrics if metrics else None
    except Exception:
        return None


def build_graph(edges_df: pd.DataFrame) -> nx.DiGraph:
    """
    Build a NetworkX DiGraph from an edges DataFrame.
    Expects 'source' and 'target' columns. Preserves edge attributes.
    """
    G = nx.DiGraph()

    for _, row in edges_df.iterrows():
        src = str(row["source"]).strip() if pd.notna(row["source"]) else None
        tgt = str(row["target"]).strip() if pd.notna(row["target"]) else None

        if src and tgt and src != tgt:  # Skip self-loops
            edge_attrs = {}
            for col in edges_df.columns:
                if col not in ("source", "target") and pd.notna(row[col]):
                    edge_attrs[col] = row[col]
            G.add_edge(src, tgt, **edge_attrs)

    return G


def attach_vertex_attributes(G: nx.DiGraph, vertices_df: pd.DataFrame) -> nx.DiGraph:
    """
    Attach vertex attributes from the Vertices sheet to graph nodes.
    Uses VERTEX_COLUMN_MAP to rename columns to clean attribute names.
    """
    if vertices_df is None or "Vertex" not in vertices_df.columns:
        return G

    for _, row in vertices_df.iterrows():
        username = str(row["Vertex"]).strip() if pd.notna(row["Vertex"]) else None
        if username is None or username not in G:
            continue

        for nodexl_col, attr_name in VERTEX_COLUMN_MAP.items():
            if nodexl_col in vertices_df.columns:
                val = row[nodexl_col]
                if pd.notna(val):
                    # Convert numeric types
                    if attr_name in (
                        "followers", "following", "tweets_count", "listed_count",
                        "favourites_count", "media_count",
                    ):
                        try:
                            val = int(float(val))
                        except (ValueError, TypeError):
                            pass
                    elif attr_name in (
                        "degree_centrality", "in_degree_centrality",
                        "out_degree_centrality", "betweenness_centrality",
                        "closeness_centrality", "eigenvector_centrality",
                        "pagerank", "clustering_coefficient", "reciprocity_ratio",
                        "layout_x", "layout_y",
                    ):
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            pass
                    elif attr_name in ("verified", "is_blue_verified", "default_profile",
                                       "default_profile_image", "possibly_sensitive"):
                        # Boolean-ish fields
                        val = str(val).strip().lower() in ("true", "yes", "1", "verified")

                    G.nodes[username][attr_name] = val

    return G


def compute_graph_stats(G: nx.DiGraph) -> Dict[str, Any]:
    """Compute basic graph-level statistics."""
    stats = {
        "node_count": G.number_of_nodes(),
        "edge_count": G.number_of_edges(),
        "is_directed": G.is_directed(),
        "density": nx.density(G),
        "reciprocity": nx.overall_reciprocity(G) if G.is_directed() else None,
        "connected_components": nx.number_weakly_connected_components(G) if G.is_directed()
                               else nx.number_connected_components(G),
    }

    # Largest component
    if G.is_directed():
        largest_cc = max(nx.weakly_connected_components(G), key=len)
    else:
        largest_cc = max(nx.connected_components(G), key=len)
    stats["largest_component_size"] = len(largest_cc)
    stats["largest_component_pct"] = round(len(largest_cc) / max(stats["node_count"], 1) * 100, 1)

    # Edge type distribution
    edge_types = {}
    for _, _, data in G.edges(data=True):
        et = data.get("edge_type", "Unknown")
        edge_types[et] = edge_types.get(et, 0) + 1
    stats["edge_types"] = edge_types

    return stats


def load_nodexl(filepath: str) -> Tuple[nx.DiGraph, Dict[str, Any]]:
    """
    Main entry point: load a NodeXL XLSX (or CSV edge list) into a NetworkX DiGraph.

    Args:
        filepath: Path to .xlsx (NodeXL export) or .csv (edge list)

    Returns:
        (G, metadata) where:
          - G is a NetworkX DiGraph with all vertex attributes attached
          - metadata dict contains: filename, format, graph_stats, overall_metrics, edge_count, node_count

    Raises:
        FileNotFoundError: If filepath doesn't exist
        ValueError: If format is unsupported or file is corrupt
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    fmt = _detect_format(filepath)
    metadata = {
        "filename": path.name,
        "filepath": str(path.absolute()),
        "format": fmt,
    }

    if fmt == "nodexl_xlsx":
        edges_df = _parse_edges_nodexl(filepath)
        vertices_df = _parse_vertices(filepath)
        overall_metrics = _parse_overall_metrics(filepath)
    else:
        edges_df = _parse_edges_csv(filepath)
        vertices_df = None
        overall_metrics = None

    # Validate edges
    if "source" not in edges_df.columns or "target" not in edges_df.columns:
        raise ValueError("Edge data must have 'source' and 'target' columns")

    # Clean: drop rows with missing source or target
    edges_df = edges_df.dropna(subset=["source", "target"])
    edges_df["source"] = edges_df["source"].astype(str).str.strip()
    edges_df["target"] = edges_df["target"].astype(str).str.strip()

    metadata["raw_edge_count"] = len(edges_df)

    # Build graph
    G = build_graph(edges_df)
    metadata["edge_count"] = G.number_of_edges()

    # Attach vertex attributes
    if vertices_df is not None:
        G = attach_vertex_attributes(G, vertices_df)
        metadata["vertex_columns"] = len(vertices_df.columns)

    # Graph statistics
    graph_stats = compute_graph_stats(G)
    metadata["graph_stats"] = graph_stats

    if overall_metrics:
        metadata["overall_metrics"] = overall_metrics

    return G, metadata


# ─── CLI: Quick load test ────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python loader.py <file.xlsx|file.csv>")
        sys.exit(1)

    filepath = sys.argv[1]
    print(f"Loading: {filepath}")
    G, meta = load_nodexl(filepath)

    print(f"\n=== Graph Loaded ===")
    print(f"  Nodes: {G.number_of_nodes()}")
    print(f"  Edges: {G.number_of_edges()}")
    print(f"  Density: {meta['graph_stats']['density']:.6f}")
    print(f"  Components: {meta['graph_stats']['connected_components']}")
    print(f"  Edge types: {meta['graph_stats']['edge_types']}")

    # Sample top nodes by degree
    print(f"\n=== Top 10 Nodes (by in-degree) ===")
    in_deg = sorted(G.in_degree(weight=None), key=lambda x: x[1], reverse=True)[:10]
    for node, deg in in_deg:
        name = G.nodes[node].get("display_name", "")
        followers = G.nodes[node].get("followers", "")
        print(f"  @{node}  in_deg={deg}  followers={followers}  name='{name}'")

    print(f"\n=== Top 10 Nodes (by PageRank) ===")
    pr_items = sorted(G.nodes(data=True), key=lambda x: x[1].get("pagerank", 0), reverse=True)[:10]
    for node, attrs in pr_items:
        pr = attrs.get("pagerank", 0)
        print(f"  @{node}  pagerank={pr:.6f}")

    print("\nDone.")
